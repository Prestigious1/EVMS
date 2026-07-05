"""
EVMS Universal Report Engine
=============================
Centralised helpers for date-range filtering, advanced filtering, search,
sorting, export (CSV / XLSX / PDF / print) and audit logging used by every
report view in the system.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from django.db.models import Q, QuerySet
from django.http import HttpResponse
from django.utils import timezone


# ---------------------------------------------------------------------------
# Date Range Engine
# ---------------------------------------------------------------------------

PERIOD_CHOICES = [
    ("today",          "Today"),
    ("yesterday",      "Yesterday"),
    ("last7",          "Last 7 Days"),
    ("last30",         "Last 30 Days"),
    ("this_month",     "This Month"),
    ("last_month",     "Last Month"),
    ("this_quarter",   "This Quarter"),
    ("this_year",      "This Year"),
    ("last_year",      "Last Year"),
    ("academic",       "Academic Session"),
    ("semester",       "Semester"),
    ("custom",         "Custom Range"),
]


def apply_date_range(
    request,
    qs: QuerySet,
    field: str,
    *,
    default: str = "this_month",
) -> tuple[QuerySet, date | None, date | None, str]:
    """Apply a date-range filter to *qs* based on request GET params.

    Returns (filtered_qs, start_date, end_date, period_key).
    """
    period = (request.GET.get("period") or default).lower()
    today = timezone.localdate()
    start: date | None = None
    end: date | None = None

    if period == "today":
        start = end = today

    elif period == "yesterday":
        start = end = today - timedelta(days=1)

    elif period == "last7":
        start = today - timedelta(days=6)
        end = today

    elif period == "last30":
        start = today - timedelta(days=29)
        end = today

    elif period == "this_month":
        start = today.replace(day=1)
        end = today

    elif period == "last_month":
        first_this = today.replace(day=1)
        end = first_this - timedelta(days=1)
        start = end.replace(day=1)

    elif period == "this_quarter":
        q_month = ((today.month - 1) // 3) * 3 + 1
        start = today.replace(month=q_month, day=1)
        end = today

    elif period == "this_year":
        start = today.replace(month=1, day=1)
        end = today

    elif period == "last_year":
        start = date(today.year - 1, 1, 1)
        end = date(today.year - 1, 12, 31)

    elif period == "academic":
        from core.models import AcademicPeriod
        ap = (
            AcademicPeriod.objects
            .filter(period_type="SESSION", is_current=True)
            .first()
        )
        if ap:
            start, end = ap.start_date, ap.end_date

    elif period == "semester":
        from core.models import AcademicPeriod
        ap = (
            AcademicPeriod.objects
            .filter(period_type="SEMESTER", is_current=True)
            .first()
        )
        if ap:
            start, end = ap.start_date, ap.end_date

    elif period == "custom":
        try:
            start = datetime.strptime(
                request.GET.get("start_date") or "", "%Y-%m-%d"
            ).date()
            end = datetime.strptime(
                request.GET.get("end_date") or "", "%Y-%m-%d"
            ).date()
        except (ValueError, TypeError):
            start = end = None

    if start and end:
        qs = qs.filter(**{f"{field}__gte": start, f"{field}__lte": end})

    return qs, start, end, period


# ---------------------------------------------------------------------------
# Universal Filter Engine
# ---------------------------------------------------------------------------

def apply_booking_filters(request, qs: QuerySet) -> QuerySet:
    """Apply all booking-related filter parameters from request.GET."""
    params = request.GET

    if v := params.get("status"):
        qs = qs.filter(status=v)
    if v := params.get("case_status"):
        qs = qs.filter(case_status=v)
    if v := params.get("hall"):
        qs = qs.filter(hall_id=v)
    if v := params.get("purpose"):
        qs = qs.filter(purpose=v)
    if v := params.get("coupon_status"):
        qs = qs.filter(coupon_status=v)
    if v := params.get("faculty"):
        qs = qs.filter(hall__faculty__icontains=v)
    if v := params.get("department"):
        qs = qs.filter(user__department__icontains=v)
    if v := params.get("user_id"):
        qs = qs.filter(user_id=v)

    # Amount range
    try:
        if v := params.get("min_amount"):
            qs = qs.filter(total_cost__gte=Decimal(v))
        if v := params.get("max_amount"):
            qs = qs.filter(total_cost__lte=Decimal(v))
    except Exception:
        pass

    return qs


def apply_payment_filters(request, qs: QuerySet) -> QuerySet:
    """Apply payment-related filter parameters."""
    params = request.GET

    if v := params.get("status"):
        qs = qs.filter(status=v)
    if v := params.get("payment_method"):
        qs = qs.filter(payment_method=v)
    if v := params.get("payment_type"):
        qs = qs.filter(payment_type=v)
    if v := params.get("hall"):
        qs = qs.filter(reservation__hall_id=v)

    try:
        if v := params.get("min_amount"):
            qs = qs.filter(amount__gte=Decimal(v))
        if v := params.get("max_amount"):
            qs = qs.filter(amount__lte=Decimal(v))
    except Exception:
        pass

    return qs


def apply_damage_filters(request, qs: QuerySet) -> QuerySet:
    params = request.GET
    if v := params.get("is_paid"):
        qs = qs.filter(is_paid=v == "true")
    if v := params.get("is_forgiven"):
        qs = qs.filter(is_forgiven=v == "true")
    if v := params.get("hall"):
        qs = qs.filter(reservation__hall_id=v)
    return qs


def apply_audit_filters(request, qs: QuerySet) -> QuerySet:
    params = request.GET
    if v := params.get("role"):
        qs = qs.filter(role=v)
    if v := params.get("module"):
        qs = qs.filter(affected_module__icontains=v)
    if v := params.get("model_name"):
        qs = qs.filter(model_name__icontains=v)
    if v := params.get("action"):
        qs = qs.filter(action__icontains=v)
    if v := params.get("user_id"):
        qs = qs.filter(user_id=v)
    return qs


# ---------------------------------------------------------------------------
# Universal Search Engine
# ---------------------------------------------------------------------------

def search_bookings(qs: QuerySet, q: str) -> QuerySet:
    if not q:
        return qs
    return qs.filter(
        Q(booking_reference__icontains=q) |
        Q(event_name__icontains=q) |
        Q(hall__name__icontains=q) |
        Q(user__email__icontains=q) |
        Q(user__first_name__icontains=q) |
        Q(user__last_name__icontains=q) |
        Q(purpose__icontains=q) |
        Q(coupon_code__icontains=q)
    )


def search_payments(qs: QuerySet, q: str) -> QuerySet:
    if not q:
        return qs
    return qs.filter(
        Q(reservation__booking_reference__icontains=q) |
        Q(user__email__icontains=q) |
        Q(paystack_reference__icontains=q) |
        Q(transaction_reference__icontains=q)
    )


def search_payment_proofs(qs: QuerySet, q: str) -> QuerySet:
    if not q:
        return qs
    return qs.filter(
        Q(reservation__booking_reference__icontains=q) |
        Q(uploaded_by__email__icontains=q) |
        Q(transaction_ref__icontains=q)
    )


def search_audit_logs(qs: QuerySet, q: str) -> QuerySet:
    if not q:
        return qs
    return qs.filter(
        Q(user__email__icontains=q) |
        Q(action__icontains=q) |
        Q(model_name__icontains=q) |
        Q(object_repr__icontains=q) |
        Q(request_id__icontains=q) |
        Q(ip_address__icontains=q)
    )


def search_users(qs: QuerySet, q: str) -> QuerySet:
    if not q:
        return qs
    return qs.filter(
        Q(email__icontains=q) |
        Q(first_name__icontains=q) |
        Q(last_name__icontains=q) |
        Q(username__icontains=q) |
        Q(department__icontains=q)
    )


# ---------------------------------------------------------------------------
# Universal Sorting Engine
# ---------------------------------------------------------------------------

BOOKING_SORT_FIELDS = {
    "newest":  "-created_at",
    "oldest":  "created_at",
    "az":      "event_name",
    "za":      "-event_name",
    "highest": "-total_cost",
    "lowest":  "total_cost",
    "date_asc": "booking_date",
    "date_desc": "-booking_date",
}

PAYMENT_SORT_FIELDS = {
    "newest":  "-created_at",
    "oldest":  "created_at",
    "highest": "-amount",
    "lowest":  "amount",
}

USER_SORT_FIELDS = {
    "newest":    "-date_joined",
    "oldest":    "date_joined",
    "az":        "email",
    "za":        "-email",
    "most_active": "-reservations",
}


def apply_sort(qs: QuerySet, sort_key: str, field_map: dict, default: str = "-created_at") -> QuerySet:
    order = field_map.get(sort_key, default)
    try:
        return qs.order_by(order)
    except Exception:
        return qs.order_by(default)


# ---------------------------------------------------------------------------
# Pagination Helper
# ---------------------------------------------------------------------------

def paginate(request, qs: QuerySet, per_page: int = 25):
    from django.core.paginator import Paginator
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get("page"))
    return page_obj


# ---------------------------------------------------------------------------
# Export Engine
# ---------------------------------------------------------------------------

def export_to_csv(
    response_or_none,
    *,
    filename: str,
    headers: list[str],
    rows: list[list[Any]],
) -> HttpResponse:
    """Build a CSV HttpResponse from headers + rows."""
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(resp)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return resp


def export_to_xlsx(
    *,
    filename: str,
    sheet_title: str,
    headers: list[str],
    rows: list[list[Any]],
    report_title: str = "",
    generated_by: str = "",
) -> HttpResponse:
    """Build an XLSX HttpResponse with professional formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel max 31 chars

    # ── University Header ──────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    gold_fill   = PatternFill("solid", fgColor="C8A96A")

    ws.merge_cells("A1:H1")
    ws["A1"] = "LAGOS STATE UNIVERSITY — ELECTRONIC VENUE MANAGEMENT SYSTEM (EVMS)"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = report_title or sheet_title
    ws["A2"].font = Font(color="C8A96A", bold=True, size=11)
    ws["A2"].fill = header_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:H3")
    from django.utils import timezone as tz
    ws["A3"] = (
        f"Generated: {tz.localtime().strftime('%d %B %Y, %H:%M')}   |   "
        f"Generated by: {generated_by or 'System'}   |   "
        f"Report: {filename}"
    )
    ws["A3"].font = Font(color="A0A0A0", size=9)
    ws["A3"].fill = PatternFill("solid", fgColor="0D1425")
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 15

    ws.append([])  # spacer row 4

    # ── Column Headers ─────────────────────────────────────────────────────
    col_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=col_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="FFFFFF")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.row_dimensions[col_row].height = 18

    # ── Data Rows ──────────────────────────────────────────────────────────
    even_fill = PatternFill("solid", fgColor="F0F4F8")
    for row_idx, row_data in enumerate(rows, col_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = even_fill
        ws.row_dimensions[row_idx].height = 15

    # ── Summary Row ────────────────────────────────────────────────────────
    if rows:
        ws.append([])
        ws.append(["", f"Total Records: {len(rows)}"])

    # ── Column Widths ──────────────────────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def export_to_pdf(
    *,
    filename: str,
    report_title: str,
    headers: list[str],
    rows: list[list[Any]],
    generated_by: str = "",
    summary_lines: list[str] | None = None,
) -> HttpResponse:
    """Build a PDF report with LASU branding using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    from django.utils import timezone as tz

    LASU_DARK  = colors.HexColor("#1E3A5F")
    LASU_GOLD  = colors.HexColor("#C8A96A")
    LASU_LIGHT = colors.HexColor("#F0F4F8")

    buf = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4

    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        textColor=LASU_DARK,
        fontSize=16,
        spaceAfter=2,
        alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "ReportSub",
        parent=styles["Normal"],
        textColor=LASU_GOLD,
        fontSize=11,
        spaceAfter=2,
        alignment=TA_CENTER,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["Normal"],
        textColor=colors.gray,
        fontSize=8,
        spaceAfter=6,
        alignment=TA_CENTER,
    )
    body_style = styles["BodyText"]

    story = []

    # ── Title block ────────────────────────────────────────────────────────
    story.append(Paragraph("LAGOS STATE UNIVERSITY", title_style))
    story.append(Paragraph("Electronic Venue Management System (EVMS)", sub_style))
    story.append(Paragraph(report_title, sub_style))
    story.append(Paragraph(
        f"Generated: {tz.localtime().strftime('%d %B %Y at %H:%M')} | "
        f"By: {generated_by or 'System'}",
        meta_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=LASU_GOLD, spaceAfter=6))

    # ── Summary lines ──────────────────────────────────────────────────────
    if summary_lines:
        for line in summary_lines:
            story.append(Paragraph(line, body_style))
        story.append(Spacer(1, 4 * mm))

    # ── Data table ────────────────────────────────────────────────────────
    table_data = [headers] + [[str(v) if v is not None else "" for v in row] for row in rows]
    col_width = (doc.width) / max(len(headers), 1)

    table = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0), LASU_DARK),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING",  (0, 0), (-1, 0), 6),
        # Body
        ("FONTSIZE",    (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LASU_LIGHT]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
    ]))
    story.append(table)

    # ── Footer note ────────────────────────────────────────────────────────
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"Total Records: {len(rows)} | This report is system-generated and confidential.",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7,
                       textColor=colors.gray, alignment=TA_CENTER),
    ))

    def _on_page(canvas, doc):
        """Draw page number and footer on each page."""
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.gray)
        page_num_text = f"Page {doc.page}"
        canvas.drawRightString(page_size[0] - 15 * mm, 10 * mm, page_num_text)
        canvas.drawString(15 * mm, 10 * mm, "LASU EVMS — Confidential")
        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    buf.seek(0)

    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ---------------------------------------------------------------------------
# Role / RBAC helpers
# ---------------------------------------------------------------------------

def get_role_scope(user) -> str:
    """Return the BI scope string for a user."""
    role = getattr(user, "role", None) or ""
    if getattr(user, "is_superuser", False) or role in ("ADMIN", "STAFF"):
        return "full"
    if role == "VENTURES":
        return "ventures"
    if role == "FACILITY":
        return "facility"
    if role == "BURSARY":
        return "bursary"
    return "none"


SCOPE_ALLOWED_REPORTS = {
    "full":     [
        "bookings", "payments", "revenue", "coupons", "damage",
        "inspections", "penalties", "halls", "applicants",
        "management", "notifications", "communications", "audit", "system_usage",
    ],
    "ventures": ["bookings", "revenue", "coupons", "communications", "payments", "applicants", "notifications"],
    "facility": ["halls", "inspections", "damage", "bookings", "notifications"],
    "bursary":  ["payments", "revenue", "damage", "penalties", "notifications"],
}


def can_access_report(user, report_key: str) -> bool:
    scope = get_role_scope(user)
    return report_key in SCOPE_ALLOWED_REPORTS.get(scope, [])
